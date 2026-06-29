from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import tomllib
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


CANONICAL_COLUMNS: tuple[str, ...] = (
    "case_id",
    "paper_case",
    "baseline_method",
    "baseline_upper_bound",
    "baseline_elapsed_s",
    "source_file",
    "source_sheet",
    "source_row",
)

MANIFEST_VERSION = 1
SUITE = "qec_code_random_window"
REQUIRED_COLUMNS = (
    "paper_case",
    "baseline_method",
    "baseline_upper_bound",
    "baseline_elapsed_s",
)
REQUIRED_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "paper_case": ("paper_case", "case", "name", "code", "dataset", "code_name", "label"),
    "baseline_method": ("baseline_method", "method", "algorithm", "decoder"),
    "baseline_upper_bound": (
        "baseline_upper_bound",
        "upper_bound",
        "ub",
        "distance",
        "d",
    ),
    "baseline_elapsed_s": (
        "baseline_elapsed_s",
        "elapsed_s",
        "seconds",
        "time_s",
        "runtime_s",
        "runtime",
        "elapsed",
        "time",
        "wall_time",
        "walltime",
    ),
}
METHOD_SHEET_ALIASES = {
    "qdistrndmw": "QDistRndMW",
    "qdist_rnd_mw": "QDistRndMW",
    "qdistevol": "QDistEvol",
    "qdist_evol": "QDistEvol",
}
SELECTED_NAME_TOKENS = ("bb", "bivariate", "qc", "summary")
BASELINE_KEY_TO_PAPER_CASE = {
    "codeDistancePYPI:bivariate_bicycle:bb72": "bb72",
    "codeDistancePYPI:bivariate_bicycle:bb144": "bb144",
}
PAPER_CASE_ALIASES: dict[str, tuple[str, ...]] = {
    "bb72": ("bb72", "bb 72", "bb-72", "[[72,12,6]]", "72,12,6"),
    "bb144": ("bb144", "bb 144", "bb-144", "[[144,12,12]]", "144,12,12"),
}
WORKBOOK_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL_NS = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
DOC_REL_NS = {
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
}


@dataclass(frozen=True)
class SheetRow:
    source_file: str
    source_sheet: str
    source_row: int
    values: dict[str, str]


def load_manifest(path: Path) -> dict[str, Any]:
    with path.open("rb") as handle:
        manifest = tomllib.load(handle)
    if not isinstance(manifest, dict):
        raise ValueError("manifest root must be a TOML table")
    if manifest.get("manifest_version") != MANIFEST_VERSION:
        raise ValueError("manifest_version must be 1")
    if manifest.get("suite") != SUITE:
        raise ValueError(f'suite must be "{SUITE}"')
    return manifest


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _parse_manifest_cases(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    cases = manifest.get("cases")
    if not isinstance(cases, list):
        raise ValueError('manifest field "cases" must be an array')
    for index, case in enumerate(cases):
        if not isinstance(case, dict):
            raise ValueError(f"case[{index}] must be a TOML table")
    return cases


def _required_case_aliases(cases: list[dict[str, Any]]) -> tuple[dict[str, tuple[int, str]], list[str]]:
    aliases: dict[str, tuple[int, str]] = {}
    required_missing: list[str] = []
    for index, case in enumerate(cases):
        case_id = case.get("case_id")
        baseline_key = case.get("baseline_key")
        baseline_required = case.get("baseline_required")
        if not isinstance(case_id, str):
            raise ValueError(f'case[{index}] field "case_id" must be a string')
        if baseline_required is True:
            paper_case = BASELINE_KEY_TO_PAPER_CASE.get(baseline_key)
            if paper_case is None:
                required_missing.append(case_id)
                continue
            aliases[paper_case] = (index, case_id)
    return aliases, required_missing


def _sheet_rows_with_openpyxl(path: Path) -> dict[str, list[list[str]]]:
    import openpyxl  # type: ignore

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        rows_by_sheet: dict[str, list[list[str]]] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_by_sheet[sheet_name] = [
                [_stringify_cell(value) for value in row]
                for row in sheet.iter_rows(values_only=True)
            ]
        return rows_by_sheet
    finally:
        workbook.close()


def _load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        xml_data = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ElementTree.fromstring(xml_data)
    strings: list[str] = []
    for string_item in root.findall("main:si", WORKBOOK_NS):
        text = "".join(text_node.text or "" for text_node in string_item.iterfind(".//main:t", WORKBOOK_NS))
        strings.append(text)
    return strings


def _column_index(ref: str) -> int:
    letters = "".join(ch for ch in ref if ch.isalpha())
    index = 0
    for ch in letters:
        index = index * 26 + (ord(ch.upper()) - ord("A") + 1)
    return index - 1


def _cell_value(cell: ElementTree.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text_node.text or "" for text_node in cell.iterfind(".//main:t", WORKBOOK_NS))
    value_node = cell.find("main:v", WORKBOOK_NS)
    if value_node is None or value_node.text is None:
        return ""
    raw = value_node.text
    if cell_type == "s":
        return shared_strings[int(raw)]
    return raw


def _sheet_name_to_target(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    relationships_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in relationships_root.findall("rel:Relationship", REL_NS)
    }
    targets: dict[str, str] = {}
    for sheet in workbook_root.findall("main:sheets/main:sheet", WORKBOOK_NS):
        name = sheet.attrib["name"]
        rel_id = sheet.attrib[f"{{{DOC_REL_NS['rel']}}}id"]
        target = rel_targets[rel_id]
        if not target.startswith("xl/"):
            target = f"xl/{target.lstrip('/')}"
        targets[name] = target
    return targets


def _sheet_rows_with_stdlib(path: Path) -> dict[str, list[list[str]]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = _load_shared_strings(archive)
        targets = _sheet_name_to_target(archive)
        rows_by_sheet: dict[str, list[list[str]]] = {}
        for sheet_name, target in targets.items():
            root = ElementTree.fromstring(archive.read(target))
            parsed_rows: list[list[str]] = []
            for row in root.findall("main:sheetData/main:row", WORKBOOK_NS):
                values_by_index: dict[int, str] = {}
                max_index = -1
                for cell in row.findall("main:c", WORKBOOK_NS):
                    ref = cell.attrib.get("r")
                    if ref is None:
                        continue
                    index = _column_index(ref)
                    values_by_index[index] = _cell_value(cell, shared_strings)
                    max_index = max(max_index, index)
                if max_index < 0:
                    parsed_rows.append([])
                    continue
                parsed_rows.append(
                    [values_by_index.get(index, "") for index in range(max_index + 1)]
                )
            rows_by_sheet[sheet_name] = parsed_rows
        return rows_by_sheet


def workbook_rows(path: Path) -> dict[str, list[list[str]]]:
    try:
        return _sheet_rows_with_openpyxl(path)
    except ModuleNotFoundError:
        return _sheet_rows_with_stdlib(path)


def _normalize_name(value: str) -> str:
    normalized = re.sub(r"[^0-9a-zA-Z]+", "_", value.strip().lower()).strip("_")
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized


def _is_selected_name(value: str) -> bool:
    normalized = _normalize_name(value)
    return any(token in normalized for token in SELECTED_NAME_TOKENS)


def _required_sheet_error() -> ValueError:
    tokens = ", ".join(SELECTED_NAME_TOKENS)
    return ValueError(
        "missing required sheet in selected workbooks; expected sheet name containing one of: "
        + tokens
    )


def _paper_case_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for canonical, aliases in PAPER_CASE_ALIASES.items():
        for alias in aliases:
            lookup[_normalize_name(alias)] = canonical
    return lookup


PAPER_CASE_LOOKUP = _paper_case_lookup()


def _method_from_sheet_name(sheet_name: str) -> str | None:
    normalized = _normalize_name(sheet_name)
    for alias, canonical in METHOD_SHEET_ALIASES.items():
        if alias in normalized:
            return canonical
    return None


def _header_indexes(row: list[str]) -> dict[str, int]:
    return {_normalize_name(value): index for index, value in enumerate(row) if value.strip()}


def _resolve_column_indexes(
    header_indexes: dict[str, int], *, method_from_sheet: str | None
) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for column in REQUIRED_COLUMNS:
        if column == "baseline_method" and method_from_sheet is not None:
            continue
        for alias in REQUIRED_COLUMN_ALIASES[column]:
            index = header_indexes.get(_normalize_name(alias))
            if index is not None:
                indexes[column] = index
                break
    return indexes


def _missing_required_columns(
    indexes: dict[str, int], *, method_from_sheet: str | None
) -> list[str]:
    missing: list[str] = []
    for column in REQUIRED_COLUMNS:
        if column == "baseline_method" and method_from_sheet is not None:
            continue
        if column not in indexes:
            missing.append(column)
    return missing


def _first_non_empty_row(sheet_rows: list[list[str]]) -> tuple[int, list[str]] | None:
    for row_number, row in enumerate(sheet_rows, start=1):
        if any(value.strip() for value in row):
            return row_number, row
    return None


def _matched_case_column_alias(header_indexes: dict[str, int]) -> bool:
    return any(
        _normalize_name(alias) in header_indexes
        for alias in REQUIRED_COLUMN_ALIASES["paper_case"]
    )


def _extract_sheet_rows(path: Path) -> list[SheetRow]:
    rows_by_sheet = workbook_rows(path)

    extracted: list[SheetRow] = []
    extracted_sheets: set[str] = set()
    for sheet_name, sheet_rows in rows_by_sheet.items():
        method_from_sheet = _method_from_sheet_name(sheet_name)
        header_specs: list[tuple[int, dict[str, int]]] = []
        for header_row_index, header_row in enumerate(sheet_rows):
            header_indexes = _header_indexes(header_row)
            if not header_indexes:
                continue
            indexes = _resolve_column_indexes(
                header_indexes, method_from_sheet=method_from_sheet
            )
            missing = _missing_required_columns(
                indexes, method_from_sheet=method_from_sheet
            )
            if missing:
                continue
            header_specs.append((header_row_index, indexes))

        if header_specs:
            extracted_sheets.add(sheet_name)
        for spec_index, (header_row_index, indexes) in enumerate(header_specs):
            next_header_index = (
                header_specs[spec_index + 1][0]
                if spec_index + 1 < len(header_specs)
                else len(sheet_rows)
            )
            for row_index in range(header_row_index + 1, next_header_index):
                row = sheet_rows[row_index]
                values = {}
                for column in REQUIRED_COLUMNS:
                    if column == "baseline_method" and method_from_sheet is not None:
                        values[column] = method_from_sheet
                        continue
                    index = indexes[column]
                    values[column] = row[index].strip() if index < len(row) else ""
                if not any(values.values()):
                    continue
                extracted.append(
                    SheetRow(
                        source_file=path.name,
                        source_sheet=sheet_name,
                        source_row=row_index + 1,
                        values=values,
                    )
                )

    for sheet_name, sheet_rows in rows_by_sheet.items():
        if not _is_selected_name(sheet_name) or sheet_name in extracted_sheets:
            continue
        first_row = _first_non_empty_row(sheet_rows)
        if first_row is None:
            raise ValueError(f'{path.name}: sheet "{sheet_name}" is empty')
        method_from_sheet = _method_from_sheet_name(sheet_name)
        for _, row in enumerate(sheet_rows, start=1):
            header_indexes = _header_indexes(row)
            if not header_indexes or not _matched_case_column_alias(header_indexes):
                continue
            indexes = _resolve_column_indexes(
                header_indexes, method_from_sheet=method_from_sheet
            )
            missing = _missing_required_columns(indexes, method_from_sheet=method_from_sheet)
            if missing:
                raise ValueError(
                    f'{path.name}: sheet "{sheet_name}" missing required column "{missing[0]}"'
                )
            raise ValueError(
                f'{path.name}: sheet "{sheet_name}" contains a recognized header that was not imported'
            )
    return extracted


def _paper_result_files(paper_results_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in paper_results_dir.iterdir()
        if path.suffix.lower() == ".xlsx"
    )


def _row_context(row: SheetRow) -> str:
    return f'{row.source_file}: sheet "{row.source_sheet}" row {row.source_row}'


def _require_non_empty_value(row: SheetRow, column: str) -> str:
    value = row.values[column].strip()
    if not value:
        raise ValueError(f"{_row_context(row)} has empty required column {column!r}")
    return value


def _require_positive_float_value(row: SheetRow, column: str) -> str:
    value = _require_non_empty_value(row, column)
    try:
        numeric = float(value)
    except ValueError as error:
        raise ValueError(
            f"{_row_context(row)} has non-numeric required column {column!r}: {value!r}"
        ) from error
    if numeric <= 0:
        raise ValueError(
            f"{_row_context(row)} has non-positive required column {column!r}: {value!r}"
        )
    return value


def _match_rows(
    required_aliases: dict[str, tuple[int, str]],
    sheet_rows: list[SheetRow],
) -> list[dict[str, str]]:
    matched: list[tuple[int, str, str, int, str, dict[str, str]]] = []
    seen_required: set[str] = set()
    for row in sheet_rows:
        paper_case = PAPER_CASE_LOOKUP.get(_normalize_name(row.values["paper_case"]))
        if paper_case is None:
            continue
        mapping = required_aliases.get(paper_case)
        if mapping is None:
            continue
        manifest_index, case_id = mapping
        baseline_method = _require_non_empty_value(row, "baseline_method")
        baseline_upper_bound = _require_positive_float_value(row, "baseline_upper_bound")
        baseline_elapsed_s = _require_positive_float_value(row, "baseline_elapsed_s")
        seen_required.add(case_id)
        matched.append(
            (
                manifest_index,
                row.source_file,
                row.source_sheet,
                row.source_row,
                baseline_method,
                {
                    "case_id": case_id,
                    "paper_case": paper_case,
                    "baseline_method": baseline_method,
                    "baseline_upper_bound": baseline_upper_bound,
                    "baseline_elapsed_s": baseline_elapsed_s,
                    "source_file": row.source_file,
                    "source_sheet": row.source_sheet,
                    "source_row": str(row.source_row),
                },
            )
        )
    missing_required = [
        case_id
        for _, case_id in sorted(required_aliases.values(), key=lambda item: item[0])
        if case_id not in seen_required
    ]
    if missing_required:
        raise ValueError(
            "missing required paper baseline rows for case_id(s): "
            + ", ".join(missing_required)
        )
    matched.sort(key=lambda item: (item[0], item[1], item[2], item[3], item[4]))
    return [row for _, _, _, _, _, row in matched]


def import_rows(cases_path: Path, paper_results_dir: Path) -> list[dict[str, str]]:
    manifest = load_manifest(cases_path)
    cases = _parse_manifest_cases(manifest)
    required_aliases, missing_required_aliases = _required_case_aliases(cases)
    if missing_required_aliases:
        raise ValueError(
            "required cases have no importer alias: " + ", ".join(missing_required_aliases)
        )

    all_sheet_rows: list[SheetRow] = []
    for workbook_path in _paper_result_files(paper_results_dir):
        all_sheet_rows.extend(_extract_sheet_rows(workbook_path))

    if not all_sheet_rows:
        raise _required_sheet_error()

    return _match_rows(required_aliases, all_sheet_rows)


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(CANONICAL_COLUMNS), lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Import codeDistancePYPI paper baselines into canonical CSV."
    )
    parser.add_argument(
        "--cases",
        type=Path,
        default=Path("benchmarks/qec_code_random_window/cases.full.toml"),
    )
    parser.add_argument("--paper-results-dir", type=Path)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args(argv)

    paper_results_dir = args.paper_results_dir
    if paper_results_dir is None:
        raw_env = os.environ.get("CODEDISTANCE_PAPER_RESULTS_DIR")
        if raw_env is None:
            print(
                "paper results directory must be provided with --paper-results-dir or CODEDISTANCE_PAPER_RESULTS_DIR",
                file=sys.stderr,
            )
            return 1
        paper_results_dir = Path(raw_env)

    try:
        rows = import_rows(args.cases, paper_results_dir)
        write_csv(args.out, rows)
    except (OSError, tomllib.TOMLDecodeError, zipfile.BadZipFile, ElementTree.ParseError, ValueError) as error:
        print(error, file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
